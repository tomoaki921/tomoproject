import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def process_workbook(filename):
    # excelを開く
    wb = xl.load_workbook(filename)
    # シートを指定する
    sheet = wb['Sheet1']

    # 10%引きの価格を計算し、指定のセルに挿入する
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # セルの参照範囲を指定する
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # グラフを生成する
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a10')

    # 保存する
    wb.save(filename)
